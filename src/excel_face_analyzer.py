# -*- coding: utf-8 -*-
"""
Phân tích khuôn mặt từ các file Excel đã tách theo từng người.
Đọc bảng "CHI TIẾT CHẤM CÔNG", tìm ngày vắng/thiếu giờ,
match ảnh trong input_images/<ngày>/ và xuất Word.
"""

import os
import re
import unicodedata
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

import xlrd
from openpyxl import load_workbook

from src.excel_extractor import ExcelToWordExporter


def _normalize_text(text: str) -> str:
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = re.sub(r'\s+', ' ', text.lower().strip())
    return text


class _ExcelReader:
    def __init__(self, path: str):
        self.path = path
        self.ext = os.path.splitext(path)[1].lower()
        self.rows: List[List] = []
        self.datemode = 0
        self._load()

    def _load(self):
        if self.ext == '.xlsx':
            wb = load_workbook(self.path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                self.rows.append([v if v is not None else '' for v in row])
        else:
            wb = xlrd.open_workbook(self.path)
            self.datemode = wb.datemode
            sheet = wb.sheet_by_index(0)
            for r in range(sheet.nrows):
                row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                self.rows.append(row)


class ExcelPersonFileParser:
    def __init__(self, path: str):
        self.path = path
        self.reader = _ExcelReader(path)
        self.header_row_idx = None
        self.col_map = {}
        self._date_formats = ["%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"]
        self._detect_header()
        self._detect_date_order()

    def _detect_header(self):
        rows = self.reader.rows
        best_row = None
        best_score = -1
        best_map = None

        for r in range(min(40, len(rows))):
            row = rows[r]
            col_map = {}
            score = 0
            for c, val in enumerate(row):
                norm = _normalize_text(val)
                if not norm:
                    continue
                if 'ma nhan vien' in norm or 'ma the' in norm or norm == 'id':
                    col_map['id'] = c
                elif 'ten nhan vien' in norm or 'ho va ten' in norm or norm == 'ten':
                    col_map['name'] = c
                elif norm == 'phong ban':
                    col_map['dept'] = c
                elif norm == 'ngay' or 'ngay' in norm:
                    col_map['date'] = c
                elif norm == 'thu':
                    col_map['weekday'] = c
                elif 'gio vao' in norm or 'check in' in norm or 'time in' in norm:
                    col_map['gio_vao'] = c
                elif 'gio ra' in norm or 'check out' in norm or 'time out' in norm:
                    col_map['gio_ra'] = c

            for k in ('id', 'name', 'date'):
                if k in col_map:
                    score += 1

            if score > best_score:
                best_score = score
                best_row = r
                best_map = col_map

        if best_row is None or best_score < 2:
            return

        self.header_row_idx = best_row
        self.col_map = best_map

    def _detect_date_order(self):
        if self.header_row_idx is None or 'date' not in self.col_map:
            return
        date_col = self.col_map['date']
        month_first = False
        day_first = False
        rows = self.reader.rows[self.header_row_idx + 1:]
        for row in rows[:200]:
            if date_col >= len(row):
                continue
            val = row[date_col]
            if not isinstance(val, str):
                continue
            s = val.strip()
            if '/' not in s and '-' not in s:
                continue
            sep = '/' if '/' in s else '-'
            parts = s.split(sep)
            if len(parts) < 3:
                continue
            try:
                p1 = int(parts[0])
                p2 = int(parts[1])
            except Exception:
                continue
            if p1 > 12 and p2 <= 12:
                day_first = True
            if p2 > 12 and p1 <= 12:
                month_first = True
            if day_first or month_first:
                break

        if month_first and not day_first:
            self._date_formats = ["%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"]
        elif day_first and not month_first:
            self._date_formats = ["%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"]

    def _get_cell(self, row: List, key: str) -> str:
        idx = self.col_map.get(key)
        if idx is None or idx >= len(row):
            return ''
        return str(row[idx]).strip()

    def _parse_date(self, val) -> Optional[date]:
        if val is None or val == '':
            return None

        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, (int, float)):
            # Heuristic: treat large floats as Excel date
            if val > 20000:
                try:
                    dt = xlrd.xldate_as_tuple(float(val), self.reader.datemode)
                    return date(dt[0], dt[1], dt[2])
                except Exception:
                    pass

        day_text = str(val).strip()
        for fmt in self._date_formats:
            try:
                return datetime.strptime(day_text, fmt).date()
            except Exception:
                pass
        return None

    def parse_person(self) -> Optional[Dict]:
        if self.header_row_idx is None:
            return None

        rows = self.reader.rows[self.header_row_idx + 1:]
        records = []
        person_name = ''
        person_id = ''
        month = None
        year = None

        for row in rows:
            date_val = row[self.col_map.get('date', 0)] if self.col_map.get('date') is not None else ''
            rec_date = self._parse_date(date_val)
            if not rec_date:
                continue

            name = self._get_cell(row, 'name')
            if name:
                person_name = name
            emp_id = self._get_cell(row, 'id')
            if emp_id:
                person_id = emp_id

            gio_vao = self._get_cell(row, 'gio_vao')
            gio_ra = self._get_cell(row, 'gio_ra')
            is_absent = not gio_vao and not gio_ra
            missing_checkout = bool(gio_vao and not gio_ra)
            missing_checkin = bool(not gio_vao and gio_ra)

            month = rec_date.month
            year = rec_date.year

            records.append({
                'day': rec_date.day,
                'date': rec_date.strftime('%d/%m/%Y'),
                'weekday': self._get_cell(row, 'weekday'),
                'gio_vao': gio_vao,
                'gio_ra': gio_ra,
                'is_absent': is_absent,
                'missing_checkout': missing_checkout,
                'missing_checkin': missing_checkin,
            })

        if not person_name or not records:
            return None

        return {
            'id': person_id,
            'name': person_name,
            'month': month or datetime.now().month,
            'year': year or datetime.now().year,
            'records': records,
        }


class ExcelFaceAnalyzer:
    def __init__(
        self,
        portrait_dir: str,
        input_images_dir: str,
        matcher,
        accuracy_mode: bool = True,
        match_distance_threshold: Optional[float] = None,
        log_detail: bool = False
    ):
        self.portrait_dir = portrait_dir
        self.input_images_dir = input_images_dir
        self.matcher = matcher
        self.accuracy_mode = accuracy_mode
        self.match_distance_threshold = match_distance_threshold
        self.log_detail = log_detail

    def analyze_folder(self, input_dir: str, output_dir: str, log_callback=None) -> List[str]:
        os.makedirs(output_dir, exist_ok=True)

        # Parse all person excel files
        persons = []
        for f in os.listdir(input_dir):
            if not f.lower().endswith(('.xlsx', '.xls')):
                continue
            path = os.path.join(input_dir, f)
            parser = ExcelPersonFileParser(path)
            person = parser.parse_person()
            if person:
                persons.append(person)
            elif log_callback:
                log_callback(f"⚠️ Không đọc được dữ liệu từ {f}", "warning")

        # Deduplicate by name: keep the file with more present rows
        picked = {}
        for person in persons:
            key = _normalize_text(person['name'])
            present = sum(1 for r in person['records'] if r['gio_vao'] or r['gio_ra'])
            score = (present, len(person['records']))
            if key not in picked or score > picked[key][0]:
                picked[key] = (score, person)

        final_persons = [v[1] for v in picked.values()]
        if log_callback:
            log_callback(f"📌 Tổng số người sẽ xử lý: {len(final_persons)}", "info")

        exporter = ExcelToWordExporter(
            self.portrait_dir,
            output_dir,
            input_images_dir=self.input_images_dir,
            face_matcher=self.matcher,
            accuracy_mode=self.accuracy_mode,
            match_distance_threshold=self.match_distance_threshold,
            log_detail=self.log_detail
        )

        results = []
        for person in final_persons:
            try:
                if log_callback:
                    issue_days = sum(
                        1 for r in person['records']
                        if r.get('is_absent') or r.get('missing_checkout') or r.get('missing_checkin')
                    )
                    log_callback(f"👤 {person['name']}: {issue_days} ngày cần đối chiếu ảnh", "default")
                path = exporter.export_person(person, log_callback=log_callback)
                results.append(path)
            except Exception:
                if log_callback:
                    log_callback(f"❌ Lỗi xuất {person['name']}", "error")
        return results
