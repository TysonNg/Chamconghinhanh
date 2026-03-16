# -*- coding: utf-8 -*-
"""
Module tách Excel chấm công dạng "CHI TIẾT CHẤM CÔNG"
Mỗi nhân viên -> 1 file Excel riêng, giữ nguyên header và cột.
"""

import os
import re
import unicodedata
from typing import Dict, List, Optional, Tuple

import xlrd
from openpyxl import Workbook, load_workbook


def _normalize_text(text: str) -> str:
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = re.sub(r'\s+', ' ', text.lower().strip())
    return text


class _ExcelReader:
    def __init__(self, path: str):
        self.path = path
        self.ext = os.path.splitext(path)[1].lower()
        self.rows: List[List] = []
        self._load()

    def _load(self):
        if self.ext == '.xlsx':
            wb = load_workbook(self.path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                self.rows.append([v if v is not None else '' for v in row])
        else:
            wb = xlrd.open_workbook(self.path)
            sheet = wb.sheet_by_index(0)
            for r in range(sheet.nrows):
                row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                self.rows.append(row)


class ExcelAttendanceSplitter:
    """
    Tách Excel chấm công dạng bảng list có header:
    STT | Mã nhân viên | Tên nhân viên | Phòng ban | Ngày | Thứ | Giờ vào | Giờ ra | ...
    """

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.reader = _ExcelReader(excel_path)
        self.header_row_idx = None
        self.title_row_idx = None
        self.col_map = {}
        self._detect_header()

    def _detect_header(self):
        rows = self.reader.rows
        best_row = None
        best_score = -1
        best_map = None

        for r in range(min(40, len(rows))):
            row = rows[r]
            col_map = {}
            score = 0
            for c, val in enumerate(row):
                norm = _normalize_text(val)
                if not norm:
                    continue
                if 'ma nhan vien' in norm or 'ma the' in norm or norm == 'id':
                    col_map['id'] = c
                elif 'ten nhan vien' in norm or 'ho va ten' in norm or norm == 'ten':
                    col_map['name'] = c
                elif norm == 'phong ban':
                    col_map['dept'] = c
                elif norm == 'ngay' or 'ngay' in norm:
                    col_map['date'] = c
                elif norm == 'thu':
                    col_map['weekday'] = c
                elif 'gio vao' in norm or 'check in' in norm or 'time in' in norm:
                    col_map['gio_vao'] = c
                elif 'gio ra' in norm or 'check out' in norm or 'time out' in norm:
                    col_map['gio_ra'] = c

            for k in ('id', 'name', 'date'):
                if k in col_map:
                    score += 1

            if score > best_score:
                best_score = score
                best_row = r
                best_map = col_map

        if best_row is None or best_score < 2:
            return

        # Detect title row above (contains "CHI TIẾT CHẤM CÔNG")
        title_row = None
        for r in range(max(0, best_row - 3), best_row + 1):
            row_text = ' '.join(_normalize_text(v) for v in rows[r] if v)
            if 'chi tiet cham cong' in row_text:
                title_row = r
                break

        self.header_row_idx = best_row
        self.title_row_idx = title_row
        self.col_map = best_map

    def _iter_data_rows(self) -> List[List]:
        rows = self.reader.rows
        if self.header_row_idx is None:
            return []
        return rows[self.header_row_idx + 1 :]

    def _row_has_data(self, row: List) -> bool:
        if not row:
            return False
        for v in row:
            if str(v).strip() != '':
                return True
        return False

    def _get_cell(self, row: List, key: str) -> str:
        idx = self.col_map.get(key)
        if idx is None or idx >= len(row):
            return ''
        return str(row[idx]).strip()

    def split(self, output_dir: str) -> Tuple[List[str], List[Dict]]:
        if self.header_row_idx is None:
            raise ValueError("Không nhận diện được header của bảng chấm công.")

        os.makedirs(output_dir, exist_ok=True)

        rows = self._iter_data_rows()
        groups: Dict[Tuple[str, str], List[List]] = {}

        for row in rows:
            if not self._row_has_data(row):
                continue
            name = self._get_cell(row, 'name')
            if not name:
                continue
            emp_id = self._get_cell(row, 'id')
            key = (_normalize_text(name), str(emp_id).strip())
            groups.setdefault(key, []).append(row)

        # Resolve duplicate names: pick group with more check-in/out rows
        selected: Dict[str, Tuple[str, List[List]]] = {}
        for (norm_name, emp_id), data_rows in groups.items():
            present_rows = 0
            for r in data_rows:
                gio_vao = self._get_cell(r, 'gio_vao')
                gio_ra = self._get_cell(r, 'gio_ra')
                if gio_vao or gio_ra:
                    present_rows += 1
            score = (present_rows, len(data_rows))

            if norm_name not in selected:
                selected[norm_name] = (emp_id, data_rows, score)
            else:
                _, _, prev_score = selected[norm_name]
                if score > prev_score:
                    selected[norm_name] = (emp_id, data_rows, score)

        header_rows = []
        if self.title_row_idx is not None and self.title_row_idx < self.header_row_idx:
            for r in range(self.title_row_idx, self.header_row_idx + 1):
                header_rows.append(self.reader.rows[r])
        else:
            header_rows.append(self.reader.rows[self.header_row_idx])

        output_files = []
        summaries = []
        for norm_name, (emp_id, data_rows, score) in selected.items():
            display_name = data_rows[0][self.col_map.get('name', 0)]
            safe_name = re.sub(r'[<>:"/\\\\|?*]', '_', str(display_name).strip())
            if emp_id:
                filename = f"{safe_name}_{emp_id}.xlsx"
            else:
                filename = f"{safe_name}.xlsx"
            out_path = os.path.join(output_dir, filename)

            wb = Workbook()
            ws = wb.active
            # Write header rows
            for hr in header_rows:
                ws.append(list(hr))
            # Write data rows for this person
            for r in data_rows:
                ws.append(list(r))
            wb.save(out_path)

            output_files.append(out_path)
            summaries.append({
                'name': str(display_name),
                'id': str(emp_id),
                'rows': len(data_rows),
                'present_rows': score[0],
            })

        return output_files, summaries

